import os
import math
import pandas as pd
import streamlit as st
import psycopg2
from openpyxl import load_workbook
from dotenv import load_dotenv
from google.cloud import vision
import re
import base64
import cv2
import numpy as np
from PIL import Image
import io
import tempfile

# Ø¥Ø¶Ø§ÙØ§Øª Ù„Ø§Ø²Ù…Ø© Ù„Ù„ØªØ¨ÙˆÙŠØ¨ Ø§Ù„Ø°ÙƒÙŠ
from rapidfuzz import process, fuzz
import time
import openpyxl

# ---- Ø§Ù„Ø¥Ø¹Ø¯Ø§Ø¯Ø§Øª Ø§Ù„Ø¹Ø§Ù…Ø© / Ø§Ù„Ø¨ÙŠØ¦Ø© ----
load_dotenv()

USERNAME = "admin"
PASSWORD = "Moraqip@123"

st.set_page_config(page_title="Ø§Ù„Ù…Ø±Ø§Ù‚Ø¨ Ø§Ù„Ø°ÙƒÙŠ", layout="wide")

# ---- Ø¥Ø¹Ø¯Ø§Ø¯ Google Vision Ù…Ù† secrets ----
def setup_google_vision():
    try:
        key_b64 = st.secrets["GOOGLE_VISION_KEY_B64"]
        key_bytes = base64.b64decode(key_b64)
        with open("google_vision.json", "wb") as f:
            f.write(key_bytes)
        os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = "google_vision.json"
        return vision.ImageAnnotatorClient()
    except Exception as e:
        st.error(f"âŒ Ù„Ù… ÙŠØªÙ… ØªØ­Ù…ÙŠÙ„ Ù…ÙØªØ§Ø­ Google Vision Ø¨Ø´ÙƒÙ„ ØµØ­ÙŠØ­: {e}")
        return None

# ---- Ø§ØªØµØ§Ù„ Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª ----
def get_conn():
    return psycopg2.connect(
        dbname=os.environ.get("DB_NAME"),
        user=os.environ.get("DB_USER"),
        password=os.environ.get("DB_PASSWORD"),
        host=os.environ.get("DB_HOST"),
        port=os.environ.get("DB_PORT"),
        sslmode=os.environ.get("DB_SSLMODE", "require")
    )

# ---- Ø¯Ø§Ù„Ø© ØªØ­ÙˆÙŠÙ„ Ø§Ù„Ø¬Ù†Ø³ ----
def map_gender(x):
    try:
        val = int(float(x))
        return "F" if val == 1 else "M"
    except:
        return "M"

# ---- ØªØ³Ø¬ÙŠÙ„ Ø§Ù„Ø¯Ø®ÙˆÙ„ ----
def login():
    st.markdown(
        """
        <style>
        .login-container {
            display: flex;
            justify-content: center;
            align-items: flex-start; /* ÙŠØ±ÙØ¹ Ø§Ù„ØµÙ†Ø¯ÙˆÙ‚ Ù„ÙÙˆÙ‚ */
            height: 100vh;
            padding-top: 10vh;       /* Ù…Ø³Ø§ÙØ© Ù…Ù† ÙÙˆÙ‚ */
        }
        .login-box {
            background: #ffffff;
            padding: 1.5rem 2rem;
            border-radius: 12px;
            box-shadow: 0px 2px 12px rgba(0,0,0,0.1);
            text-align: center;
            width: 300px;
        }
        .stTextInput>div>div>input {
            text-align: center;
            font-size: 14px;
            height: 35px;
        }
        .stButton button {
            background: linear-gradient(90deg, #4e73df, #1cc88a);
            color: white;
            border-radius: 6px;
            padding: 0.4rem 0.8rem;
            font-size: 14px;
            font-weight: bold;
            transition: 0.2s;
            width: 100%;
        }
        .stButton button:hover {
            background: linear-gradient(90deg, #1cc88a, #4e73df);
            transform: scale(1.02);
        }
        </style>
        """,
        unsafe_allow_html=True
    )

    st.markdown('<div class="login-container"><div class="login-box">', unsafe_allow_html=True)

    st.markdown("### ğŸ”‘ ØªØ³Ø¬ÙŠÙ„ Ø§Ù„Ø¯Ø®ÙˆÙ„")
    u = st.text_input("ğŸ‘¤ Ø§Ø³Ù… Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù…", key="login_user")
    p = st.text_input("ğŸ”’ ÙƒÙ„Ù…Ø© Ø§Ù„Ù…Ø±ÙˆØ±", type="password", key="login_pass")

    # âœ… ÙƒØ¨Ø³Ø© ÙˆØ§Ø­Ø¯Ø© ØªÙƒÙÙŠ
    login_btn = st.button("ğŸš€ Ø¯Ø®ÙˆÙ„", key="login_btn")
    if login_btn:
        if u == USERNAME and p == PASSWORD:
            st.session_state.logged_in = True
            st.rerun()   # Ø¥Ø¹Ø§Ø¯Ø© ØªØ­Ù…ÙŠÙ„ Ø§Ù„ØµÙØ­Ø© Ù…Ø¨Ø§Ø´Ø±Ø©
        else:
            st.error("âŒ Ø§Ø³Ù… Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù… Ø£Ùˆ ÙƒÙ„Ù…Ø© Ø§Ù„Ù…Ø±ÙˆØ± ØºÙŠØ± ØµØ­ÙŠØ­Ø©")

    st.markdown('</div></div>', unsafe_allow_html=True)


# ---- ØªØ­Ù‚Ù‚ Ù…Ù† Ø­Ø§Ù„Ø© Ø§Ù„Ø¬Ù„Ø³Ø© ----
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

if not st.session_state.logged_in:
    login()
    st.stop()

# ========================== Ø§Ù„ÙˆØ§Ø¬Ù‡Ø© Ø¨Ø¹Ø¯ ØªØ³Ø¬ÙŠÙ„ Ø§Ù„Ø¯Ø®ÙˆÙ„ ==========================
st.title("ğŸ“Š Ø¨ØºØ¯Ø§Ø¯ - Ø§Ù„Ø¨Ø­Ø« ÙÙŠ Ø³Ø¬Ù„Ø§Øª Ø§Ù„Ù†Ø§Ø®Ø¨ÙŠÙ†")
st.markdown("Ø³ÙŠØªÙ… Ø§Ù„Ø¨Ø­Ø« ÙÙŠ Ù‚ÙˆØ§Ø¹Ø¯ Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª Ø¨Ø§Ø³ØªØ®Ø¯Ø§Ù… Ø§Ù„Ø°ÙƒØ§Ø¡ Ø§Ù„Ø§ØµØ·Ù†Ø§Ø¹ÙŠ ğŸ¤–")

# ====== ØªØ¨ÙˆÙŠØ¨Ø§Øª ======
tab_browse, tab_single, tab_file, tab_file_name_center, tab_count, tab_check, tab_count_custom = st.tabs(
    [
        "ğŸ“„ ØªØµÙÙ‘Ø­ Ø§Ù„Ø³Ø¬Ù„Ø§Øª",
        "ğŸ” Ø¨Ø­Ø« Ø¨Ø±Ù‚Ù…",
        "ğŸ“‚ Ø±ÙØ¹ Ù…Ù„Ù Excel",
        "ğŸ” Ø§Ù„Ø¨Ø­Ø« Ø§Ù„Ø°ÙƒÙŠ (Ø§Ø³Ù… + Ù…Ø±ÙƒØ² Ø§Ù‚ØªØ±Ø§Ø¹)",  # ğŸ‘ˆ Ø§Ù„ØªØ¨ÙˆÙŠØ¨ Ø§Ù„Ø±Ø§Ø¨Ø¹ Ø§Ù„Ø¬Ø¯ÙŠØ¯
        "ğŸ“¦ Ø¹Ø¯Ù‘ Ø§Ù„Ø¨Ø·Ø§Ù‚Ø§Øª",
        "ğŸ§¾ Ø§Ù„ØªØ­Ù‚Ù‚ Ù…Ù† Ø§Ù„Ù…Ø¹Ù„ÙˆÙ…Ø§Øª",
        "ğŸ§® ØªØ­Ù„ÙŠÙ„ Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª (COUNT)"
    ]
)

# ----------------------------------------------------------------------------- #
# 1) ğŸ“„ ØªØµÙÙ‘Ø­ Ø§Ù„Ø³Ø¬Ù„Ø§Øª
# ----------------------------------------------------------------------------- #
with tab_browse:
    st.subheader("ğŸ“„ ØªØµÙÙ‘Ø­ Ø§Ù„Ø³Ø¬Ù„Ø§Øª Ù…Ø¹ ÙÙ„Ø§ØªØ±")

    if "page" not in st.session_state:
        st.session_state.page = 1
    if "filters" not in st.session_state:
        st.session_state.filters = {"voter": "", "name": "", "center": ""}

    colf1, colf2, colf3, colf4 = st.columns([1,1,1,1])
    with colf1:
        voter_filter = st.text_input("ğŸ”¢ Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨:", value=st.session_state.filters["voter"])
    with colf2:
        name_filter = st.text_input("ğŸ§‘â€ğŸ’¼ Ø§Ù„Ø§Ø³Ù…:", value=st.session_state.filters["name"])
    with colf3:
        center_filter = st.text_input("ğŸ« Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹:", value=st.session_state.filters["center"])
    with colf4:
        page_size = st.selectbox("Ø¹Ø¯Ø¯ Ø§Ù„ØµÙÙˆÙ", [10, 20, 50, 100], index=1)

    if st.button("ğŸ” ØªØ·Ø¨ÙŠÙ‚ Ø§Ù„ÙÙ„Ø§ØªØ±"):
        st.session_state.filters = {
            "voter": voter_filter.strip(),
            "name": name_filter.strip(),
            "center": center_filter.strip(),
        }
        st.session_state.page = 1

    # --- Ø¨Ù†Ø§Ø¡ Ø´Ø±ÙˆØ· Ø§Ù„Ø¨Ø­Ø« ---
    where_clauses, params = [], []
    if st.session_state.filters["voter"]:
        where_clauses.append('CAST("Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨" AS TEXT) ILIKE %s')
        params.append(f"%{st.session_state.filters['voter']}%")
    if st.session_state.filters["name"]:
        where_clauses.append('"Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø«Ù„Ø§Ø«ÙŠ" ILIKE %s')
        params.append(f"%{st.session_state.filters['name']}%")
    if st.session_state.filters["center"]:
        where_clauses.append('"Ø§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹" ILIKE %s')
        params.append(f"%{st.session_state.filters['center']}%")

    where_sql = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""

    count_sql = f'SELECT COUNT(*) FROM "Bagdad" {where_sql};'
    offset = (st.session_state.page - 1) * page_size
    data_sql = f'''
        SELECT
            "Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨","Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø«Ù„Ø§Ø«ÙŠ","Ø§Ù„Ø¬Ù†Ø³","Ù‡Ø§ØªÙ","Ø±Ù‚Ù… Ø§Ù„Ø¹Ø§Ø¦Ù„Ø©",
            "Ø§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹","Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹",
            "Ø§Ù„Ù…Ø¯ÙŠÙ†Ø©","Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„ØªØ³Ø¬ÙŠÙ„","Ø§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„ØªØ³Ø¬ÙŠÙ„","ØªØ§Ø±ÙŠØ® Ø§Ù„Ù…ÙŠÙ„Ø§Ø¯"
        FROM "Bagdad"
        {where_sql}
        ORDER BY "Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨" ASC
        LIMIT %s OFFSET %s;
    '''

    try:
        conn = get_conn()
        with conn.cursor() as cur:
            cur.execute(count_sql, params)
            total_rows = cur.fetchone()[0]

        df = pd.read_sql_query(data_sql, conn, params=params + [page_size, offset])
        conn.close()

        if not df.empty:
            df = df.rename(columns={
                "Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨": "Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨",
                "Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø«Ù„Ø§Ø«ÙŠ": "Ø§Ù„Ø§Ø³Ù…",
                "Ø§Ù„Ø¬Ù†Ø³": "Ø§Ù„Ø¬Ù†Ø³",
                "Ù‡Ø§ØªÙ": "Ø±Ù‚Ù… Ø§Ù„Ù‡Ø§ØªÙ",
                "Ø±Ù‚Ù… Ø§Ù„Ø¹Ø§Ø¦Ù„Ø©": "Ø±Ù‚Ù… Ø§Ù„Ø¹Ø§Ø¦Ù„Ø©",
                "Ø§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹": "Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹",
                "Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹": "Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹",
                "Ø§Ù„Ù…Ø¯ÙŠÙ†Ø©": "Ø§Ù„Ù…Ø¯ÙŠÙ†Ø©",
                "Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„ØªØ³Ø¬ÙŠÙ„": "Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„ØªØ³Ø¬ÙŠÙ„",
                "Ø§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„ØªØ³Ø¬ÙŠÙ„": "Ø§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„ØªØ³Ø¬ÙŠÙ„",
                "ØªØ§Ø±ÙŠØ® Ø§Ù„Ù…ÙŠÙ„Ø§Ø¯": "ØªØ§Ø±ÙŠØ® Ø§Ù„Ù…ÙŠÙ„Ø§Ø¯"
            })
            df["Ø§Ù„Ø¬Ù†Ø³"] = df["Ø§Ù„Ø¬Ù†Ø³"].apply(map_gender)

        total_pages = max(1, math.ceil(total_rows / page_size))

        # âœ… Ø¹Ø±Ø¶ Ø§Ù„Ù†ØªØ§Ø¦Ø¬
        st.dataframe(df, use_container_width=True, height=500)

        c1, c2, c3 = st.columns([1,2,1])
        with c1:
            if st.button("â¬…ï¸ Ø§Ù„Ø³Ø§Ø¨Ù‚", disabled=(st.session_state.page <= 1)):
                st.session_state.page -= 1
                st.experimental_rerun()
        with c2:
            st.markdown(f"<div style='text-align:center;font-weight:bold'>ØµÙØ­Ø© {st.session_state.page} Ù…Ù† {total_pages}</div>", unsafe_allow_html=True)
        with c3:
            if st.button("Ø§Ù„ØªØ§Ù„ÙŠ â¡ï¸", disabled=(st.session_state.page >= total_pages)):
                st.session_state.page += 1
                st.experimental_rerun()

    except Exception as e:
        st.error(f"âŒ Ø®Ø·Ø£ Ø£Ø«Ù†Ø§Ø¡ Ø§Ù„ØªØµÙØ­: {e}")

# ----------------------------------------------------------------------------- #
# 2) ğŸ” Ø§Ù„Ø¨Ø­Ø« Ø¨Ø±Ù‚Ù… ÙˆØ§Ø­Ø¯
# ----------------------------------------------------------------------------- #
with tab_single:
    st.subheader("ğŸ” Ø§Ù„Ø¨Ø­Ø« Ø¨Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨")
    voter_input = st.text_input("Ø§Ø¯Ø®Ù„ Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨:")
    if st.button("Ø¨Ø­Ø«"):
        try:
            conn = get_conn()
            query = """
                SELECT "Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨","Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø«Ù„Ø§Ø«ÙŠ","Ø§Ù„Ø¬Ù†Ø³","Ù‡Ø§ØªÙ","Ø±Ù‚Ù… Ø§Ù„Ø¹Ø§Ø¦Ù„Ø©",
                       "Ø§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹","Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹",
                       "Ø§Ù„Ù…Ø¯ÙŠÙ†Ø©","Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„ØªØ³Ø¬ÙŠÙ„","Ø§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„ØªØ³Ø¬ÙŠÙ„","ØªØ§Ø±ÙŠØ® Ø§Ù„Ù…ÙŠÙ„Ø§Ø¯"
                FROM "Bagdad" WHERE "Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨" LIKE %s
            """
            df = pd.read_sql_query(query, conn, params=(voter_input.strip(),))
            conn.close()

            if not df.empty:
                df = df.rename(columns={
                    "Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨": "Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨",
                    "Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø«Ù„Ø§Ø«ÙŠ": "Ø§Ù„Ø§Ø³Ù…",
                    "Ø§Ù„Ø¬Ù†Ø³": "Ø§Ù„Ø¬Ù†Ø³",
                    "Ù‡Ø§ØªÙ": "Ø±Ù‚Ù… Ø§Ù„Ù‡Ø§ØªÙ",
                    "Ø±Ù‚Ù… Ø§Ù„Ø¹Ø§Ø¦Ù„Ø©": "Ø±Ù‚Ù… Ø§Ù„Ø¹Ø§Ø¦Ù„Ø©",
                    "Ø§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹": "Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹",
                    "Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹": "Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹",
                    "Ø§Ù„Ù…Ø¯ÙŠÙ†Ø©": "Ø§Ù„Ù…Ø¯ÙŠÙ†Ø©",
                    "Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„ØªØ³Ø¬ÙŠÙ„": "Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„ØªØ³Ø¬ÙŠÙ„",
                    "Ø§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„ØªØ³Ø¬ÙŠÙ„": "Ø§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„ØªØ³Ø¬ÙŠÙ„",
                    "ØªØ§Ø±ÙŠØ® Ø§Ù„Ù…ÙŠÙ„Ø§Ø¯": "ØªØ§Ø±ÙŠØ® Ø§Ù„Ù…ÙŠÙ„Ø§Ø¯"
                })
                df["Ø§Ù„Ø¬Ù†Ø³"] = df["Ø§Ù„Ø¬Ù†Ø³"].apply(map_gender)

                st.dataframe(df, use_container_width=True, height=500)
            else:
                st.warning("âš ï¸ Ù„Ù… ÙŠØªÙ… Ø§Ù„Ø¹Ø«ÙˆØ± Ø¹Ù„Ù‰ Ù†ØªØ§Ø¦Ø¬")
        except Exception as e:
            st.error(f"âŒ Ø®Ø·Ø£: {e}")

# ----------------------------------------------------------------------------- #
# 3) ğŸ“‚ Ø±ÙØ¹ Ù…Ù„Ù Excel (Ù…Ø¹Ø¯Ù„ Ù…Ø¹ Ø§Ù„Ø£Ø±Ù‚Ø§Ù… ØºÙŠØ± Ø§Ù„Ù…ÙˆØ¬ÙˆØ¯Ø©)
# ----------------------------------------------------------------------------- #
with tab_file:
    st.subheader("ğŸ“‚ Ø§Ù„Ø¨Ø­Ø« Ø¨Ø§Ø³ØªØ®Ø¯Ø§Ù… Ù…Ù„Ù Excel")
    uploaded_file = st.file_uploader("ğŸ“¤ Ø§Ø±ÙØ¹ Ù…Ù„Ù (Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨)", type=["xlsx"])
    if uploaded_file and st.button("ğŸš€ ØªØ´ØºÙŠÙ„ Ø§Ù„Ø¨Ø­Ø«"):
        try:
            voters_df = pd.read_excel(uploaded_file, engine="openpyxl")
            voter_col = "Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨" if "Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨" in voters_df.columns else "VoterNo"
            voters_list = voters_df[voter_col].astype(str).tolist()

            conn = get_conn()
            placeholders = ",".join(["%s"] * len(voters_list))
            query = f"""
                SELECT "Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨","Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø«Ù„Ø§Ø«ÙŠ","Ø§Ù„Ø¬Ù†Ø³","Ù‡Ø§ØªÙ","Ø±Ù‚Ù… Ø§Ù„Ø¹Ø§Ø¦Ù„Ø©",
                       "Ø§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹","Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹",
                       "Ø§Ù„Ù…Ø¯ÙŠÙ†Ø©","Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„ØªØ³Ø¬ÙŠÙ„","Ø§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„ØªØ³Ø¬ÙŠÙ„","ØªØ§Ø±ÙŠØ® Ø§Ù„Ù…ÙŠÙ„Ø§Ø¯"
                FROM "Bagdad" WHERE "Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨" IN ({placeholders})
            """
            df = pd.read_sql_query(query, conn, params=voters_list)
            conn.close()

            if not df.empty:
                df = df.rename(columns={
                    "Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨": "Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨",
                    "Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø«Ù„Ø§Ø«ÙŠ": "Ø§Ù„Ø§Ø³Ù…",
                    "Ø§Ù„Ø¬Ù†Ø³": "Ø§Ù„Ø¬Ù†Ø³",
                    "Ù‡Ø§ØªÙ": "Ø±Ù‚Ù… Ø§Ù„Ù‡Ø§ØªÙ",
                    "Ø±Ù‚Ù… Ø§Ù„Ø¹Ø§Ø¦Ù„Ø©": "Ø±Ù‚Ù… Ø§Ù„Ø¹Ø§Ø¦Ù„Ø©",
                    "Ø§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹": "Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹",
                    "Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹": "Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹",
                    "Ø§Ù„Ù…Ø¯ÙŠÙ†Ø©": "Ø§Ù„Ù…Ø¯ÙŠÙ†Ø©",
                    "Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„ØªØ³Ø¬ÙŠÙ„": "Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„ØªØ³Ø¬ÙŠÙ„",
                    "Ø§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„ØªØ³Ø¬ÙŠÙ„": "Ø§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„ØªØ³Ø¬ÙŠÙ„",
                    "ØªØ§Ø±ÙŠØ® Ø§Ù„Ù…ÙŠÙ„Ø§Ø¯": "ØªØ§Ø±ÙŠØ® Ø§Ù„Ù…ÙŠÙ„Ø§Ø¯"
                })
                df["Ø§Ù„Ø¬Ù†Ø³"] = df["Ø§Ù„Ø¬Ù†Ø³"].apply(map_gender)

                df["Ø±Ù‚Ù… Ø§Ù„Ù…Ù†Ø¯ÙˆØ¨ Ø§Ù„Ø±Ø¦ÙŠØ³ÙŠ"] = ""
                df["Ø§Ù„Ø­Ø§Ù„Ø©"] = 0
                df["Ù…Ù„Ø§Ø­Ø¸Ø©"] = ""
                df["Ø±Ù‚Ù… Ø§Ù„Ù…Ø­Ø·Ø©"] = 1

                df = df[["Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨","Ø§Ù„Ø§Ø³Ù…","Ø§Ù„Ø¬Ù†Ø³","Ø±Ù‚Ù… Ø§Ù„Ù‡Ø§ØªÙ",
                         "Ø±Ù‚Ù… Ø§Ù„Ø¹Ø§Ø¦Ù„Ø©","Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹","Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹","Ø±Ù‚Ù… Ø§Ù„Ù…Ø­Ø·Ø©",
                         "Ø±Ù‚Ù… Ø§Ù„Ù…Ù†Ø¯ÙˆØ¨ Ø§Ù„Ø±Ø¦ÙŠØ³ÙŠ","Ø§Ù„Ø­Ø§Ù„Ø©","Ù…Ù„Ø§Ø­Ø¸Ø©"]]

                # âœ… Ø¥ÙŠØ¬Ø§Ø¯ Ø§Ù„Ø£Ø±Ù‚Ø§Ù… ØºÙŠØ± Ø§Ù„Ù…ÙˆØ¬ÙˆØ¯Ø©
                found_numbers = set(df["Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨"].astype(str).tolist())
                missing_numbers = [num for num in voters_list if num not in found_numbers]

                # Ø¹Ø±Ø¶ Ø§Ù„Ù†ØªØ§Ø¦Ø¬ Ø§Ù„Ù…ÙˆØ¬ÙˆØ¯Ø©
                st.dataframe(df, use_container_width=True, height=500)

                # Ù…Ù„Ù Ø§Ù„Ù†ØªØ§Ø¦Ø¬ Ø§Ù„Ù…ÙˆØ¬ÙˆØ¯Ø©
                output_file = "Ù†ØªØ§Ø¦Ø¬_Ø§Ù„Ø¨Ø­Ø«.xlsx"
                df.to_excel(output_file, index=False, engine="openpyxl")
                wb = load_workbook(output_file)
                wb.active.sheet_view.rightToLeft = True
                wb.save(output_file)
                with open(output_file, "rb") as f:
                    st.download_button("â¬‡ï¸ ØªØ­Ù…ÙŠÙ„ Ø§Ù„Ù†ØªØ§Ø¦Ø¬", f,
                        file_name="Ù†ØªØ§Ø¦Ø¬_Ø§Ù„Ø¨Ø­Ø«.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

                # Ø¹Ø±Ø¶ ÙˆØªØ­Ù…ÙŠÙ„ Ø§Ù„Ø£Ø±Ù‚Ø§Ù… ØºÙŠØ± Ø§Ù„Ù…ÙˆØ¬ÙˆØ¯Ø©
                if missing_numbers:
                    st.warning("âš ï¸ Ø§Ù„Ø£Ø±Ù‚Ø§Ù… Ø§Ù„ØªØ§Ù„ÙŠØ© Ù„Ù… ÙŠØªÙ… Ø§Ù„Ø¹Ø«ÙˆØ± Ø¹Ù„ÙŠÙ‡Ø§ ÙÙŠ Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª:")
                    st.write(missing_numbers)

                    missing_df = pd.DataFrame(missing_numbers, columns=["Ø§Ù„Ø£Ø±Ù‚Ø§Ù… ØºÙŠØ± Ø§Ù„Ù…ÙˆØ¬ÙˆØ¯Ø©"])
                    miss_file = "missing_numbers.xlsx"
                    missing_df.to_excel(miss_file, index=False, engine="openpyxl")
                    with open(miss_file, "rb") as f:
                        st.download_button("â¬‡ï¸ ØªØ­Ù…ÙŠÙ„ Ø§Ù„Ø£Ø±Ù‚Ø§Ù… ØºÙŠØ± Ø§Ù„Ù…ÙˆØ¬ÙˆØ¯Ø©", f,
                            file_name="Ø§Ù„Ø£Ø±Ù‚Ø§Ù…_ØºÙŠØ±_Ø§Ù„Ù…ÙˆØ¬ÙˆØ¯Ø©.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            else:
                st.warning("âš ï¸ Ù„Ø§ ÙŠÙˆØ¬Ø¯ Ù†ØªØ§Ø¦Ø¬")
        except Exception as e:
            st.error(f"âŒ Ø®Ø·Ø£: {e}")

# ----------------------------------------------------------------------------- #
# 4ï¸âƒ£ Ø§Ù„ØªØ¨ÙˆÙŠØ¨ Ø§Ù„Ø±Ø§Ø¨Ø¹: Ø§Ù„Ø¨Ø­Ø« Ø§Ù„Ø°ÙƒÙŠ Ø¨Ø§Ù„Ø§Ø³Ù… + Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹ (Bagdad)
# ----------------------------------------------------------------------------- #
with tab_file_name_center:
    st.subheader("ğŸ“˜ Ø§Ù„Ø¨Ø­Ø« Ø§Ù„Ø°ÙƒÙŠ (Ø¹Ø±Ø¶ Ù…Ø¨Ø§Ø´Ø± + Ø­ÙØ¸ ØªØ¯Ø±ÙŠØ¬ÙŠ ÙˆØ³Ø±ÙŠØ¹) âš¡")

    file_nc = st.file_uploader("ğŸ“¤ Ø§Ø±ÙØ¹ Ù…Ù„Ù Excel ÙŠØ­ØªÙˆÙŠ Ø§Ù„Ø§Ø³Ù… + Ø§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹", type=["xlsx"])
    run_nc = st.button("ğŸš€ Ø¨Ø¯Ø¡ Ø§Ù„Ø¨Ø­Ø« ÙˆÙ…Ø´Ø§Ù‡Ø¯Ø© Ø§Ù„ØªÙ‚Ø¯Ù…")

    # âœ… Ø¯Ø§Ù„Ø© Ø§Ù„ØªØ·Ø¨ÙŠØ¹ Ø§Ù„Ø£Ø³Ø§Ø³ÙŠØ© Ù„Ù„Ù†ØµÙˆØµ Ø§Ù„Ø¹Ø±Ø¨ÙŠØ©
    def normalize_ar(text: str) -> str:
        if not text:
            return ""
        s = str(text)
        s = s.translate(str.maketrans('', '', ''.join([
            '\u0610','\u0611','\u0612','\u0613','\u0614','\u0615','\u0616','\u0617','\u0618','\u0619','\u061A',
            '\u064B','\u064C','\u064D','\u064E','\u064F','\u0650','\u0651','\u0652','\u0653','\u0654','\u0655',
            '\u0656','\u0657','\u0658','\u0659','\u065A','\u065B','\u065C','\u065D','\u065E','\u065F','\u0670'
        ])))
        s = s.replace("Ù€", "").replace(" ", "").strip()
        s = (s.replace("Ø£","Ø§").replace("Ø¥","Ø§").replace("Ø¢","Ø§")
             .replace("Ø¤","Ùˆ").replace("Ø¦","ÙŠ").replace("Ù‰","ÙŠ").replace("Ø©","Ù‡"))
        return s.lower()

    # âœ… ØªØ·Ø¨ÙŠØ¹ Ø³Ø±ÙŠØ¹ Ø¨Ø§Ø³ØªØ®Ø¯Ø§Ù… cache
    def normalize_fast(s):
        uniq = s.fillna("").astype(str).unique()
        mapping = {u: normalize_ar(u) for u in uniq}
        return s.fillna("").astype(str).map(mapping)

    # âœ… ØªØ­Ù…ÙŠÙ„ Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„Ù…Ø±Ø§ÙƒØ² Ø¯ÙØ¹Ø© Ø¯ÙØ¹Ø© Ù…Ù† Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª (Bagdad)
    @st.cache_data(show_spinner=False)
    def load_db_for_centers(centers):
        conn = get_conn()
        all_parts = []
        batch_size = 500

        for i in range(0, len(centers), batch_size):
            batch = centers[i:i + batch_size]
            batch = [str(c).strip() for c in batch if c]

            query = """
                SELECT "Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨","Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø«Ù„Ø§Ø«ÙŠ","Ø§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹"
                FROM "Bagdad"
                WHERE CAST("Ø§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹" AS TEXT) = ANY(%(centers)s)
            """
            params = {"centers": batch}

            try:
                part = pd.read_sql_query(query, conn, params=params)
                if not part.empty:
                    all_parts.append(part)
                st.write(f"ğŸ“¥ ØªÙ… ØªØ­Ù…ÙŠÙ„ Ø¯ÙØ¹Ø© {i // batch_size + 1} ({len(batch)}) Ù…Ø±ÙƒØ²...")
            except Exception as e:
                st.warning(f"âš ï¸ Ø®Ø·Ø£ Ø£Ø«Ù†Ø§Ø¡ ØªØ­Ù…ÙŠÙ„ Ø¯ÙØ¹Ø© {i // batch_size + 1}: {e}")

        conn.close()
        return pd.concat(all_parts, ignore_index=True) if all_parts else pd.DataFrame()

    # âœ… ØªÙ†ÙÙŠØ° Ø§Ù„Ø¨Ø­Ø« Ø§Ù„ÙƒØ§Ù…Ù„
    if file_nc and run_nc:
        start = time.time()
        st.info("ğŸ“¦ Ø¬Ø§Ø±ÙŠ ØªØ¬Ù‡ÙŠØ² Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª...")

        # ---- Ù‚Ø±Ø§Ø¡Ø© Ø§Ù„Ù…Ù„Ù ----
        try:
            df = pd.read_excel(file_nc, engine="openpyxl")
            df.columns = df.columns.str.strip()
            if "Ø§Ù„Ø§Ø³Ù…" not in df.columns or "Ø§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹" not in df.columns:
                st.error("âŒ Ø§Ù„Ù…Ù„Ù ÙŠØ¬Ø¨ Ø£Ù† ÙŠØ­ØªÙˆÙŠ Ø¹Ù„Ù‰ Ø§Ù„Ø£Ø¹Ù…Ø¯Ø©: Ø§Ù„Ø§Ø³Ù… ÙˆØ§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹")
                st.stop()
        except Exception as e:
            st.error(f"âŒ Ø®Ø·Ø£ ÙÙŠ Ù‚Ø±Ø§Ø¡Ø© Ø§Ù„Ù…Ù„Ù: {e}")
            st.stop()

        # ---- ØªØ·Ø¨ÙŠØ¹ Ø§Ù„Ø£Ø³Ù…Ø§Ø¡ ÙˆØ§Ù„Ù…Ø±Ø§ÙƒØ² ----
        df["__norm_name"] = normalize_fast(df["Ø§Ù„Ø§Ø³Ù…"])
        df["__norm_center"] = normalize_fast(df["Ø§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹"])
        centers = df["Ø§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹"].dropna().unique().tolist()

        # ---- ØªØ­Ù…ÙŠÙ„ Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„Ù‚Ø§Ø¹Ø¯Ø© ----
        db_df = load_db_for_centers(centers)
        if db_df.empty:
            st.warning("âš ï¸ Ù„Ù… ÙŠØªÙ… Ø§Ù„Ø¹Ø«ÙˆØ± Ø¹Ù„Ù‰ Ø¨ÙŠØ§Ù†Ø§Øª Ù„Ù„Ù…Ø±Ø§ÙƒØ² Ø§Ù„Ù…Ø­Ø¯Ø¯Ø©.")
            st.stop()

        db_df["__norm_name"] = normalize_fast(db_df["Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø«Ù„Ø§Ø«ÙŠ"])
        db_df["__norm_center"] = normalize_fast(db_df["Ø§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹"])

        # ---- Ø¨Ù†Ø§Ø¡ ÙÙ‡Ø±Ø³ Ø³Ø±ÙŠØ¹ Ù„Ù„Ù…Ø±Ø§ÙƒØ² ----
        groups = {}
        for c, sub in db_df.groupby("__norm_center"):
            groups[c] = sub.reset_index(drop=True)

        # ---- Ù…Ù„Ù Ù…Ø¤Ù‚Øª Ù„Ù„Ø­ÙØ¸ Ø§Ù„ØªØ¯Ø±ÙŠØ¬ÙŠ ----
        temp_dir = tempfile.gettempdir()
        temp_path = os.path.join(temp_dir, "partial_results.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ù†ØªØ§Ø¦Ø¬ Ù…Ø¤Ù‚ØªØ©"
        ws.append([
            "Ø§Ù„Ø§Ø³Ù… ÙÙŠ Ø§Ù„Ù…Ù„Ù",
            "Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹ ÙÙŠ Ø§Ù„Ù…Ù„Ù",
            "Ù†Ø³Ø¨Ø© ØªØ·Ø§Ø¨Ù‚ Ø§Ù„Ø§Ø³Ù…",
            "Ù†Ø³Ø¨Ø© ØªØ·Ø§Ø¨Ù‚ Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹",
            "ÙƒÙ… ØªÙƒØ±Ø§Ø±Ù‡ ÙÙŠ Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª",
            "Ø§Ù„Ø§Ø³Ù… ÙÙŠ Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª",
            "Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨ ÙÙŠ Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª",
            "Ø§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹ ÙÙŠ Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª"
        ])
        wb.save(temp_path)

        # ---- ÙˆØ§Ø¬Ù‡Ø© Ø§Ù„ØªÙ‚Ø¯Ù… ----
        progress = st.progress(0)
        status = st.empty()
        log_box = st.empty()
        results = []
        total = len(df)

        st.info(f"ğŸš€ Ø¨Ø¯Ø¡ Ø§Ù„Ù…Ø·Ø§Ø¨Ù‚Ø©... (Ø¹Ø¯Ø¯ Ø§Ù„Ø³Ø¬Ù„Ø§Øª: {total})")

        # ---- ØªÙ†ÙÙŠØ° Ø§Ù„Ù…Ø·Ø§Ø¨Ù‚Ø© ----
        for i, row in df.iterrows():
            orig_name = str(row["Ø§Ù„Ø§Ø³Ù…"])
            orig_center = str(row["Ø§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹"])
            norm_name = row["__norm_name"]
            norm_center = row["__norm_center"]

            grp_df = groups.get(norm_center)

            # Ø§Ù„Ù‚ÙŠÙ… Ø§Ù„Ø§ÙØªØ±Ø§Ø¶ÙŠØ©
            match_row = {
                "Ø§Ù„Ø§Ø³Ù… ÙÙŠ Ø§Ù„Ù…Ù„Ù": orig_name,
                "Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹ ÙÙŠ Ø§Ù„Ù…Ù„Ù": orig_center,
                "Ù†Ø³Ø¨Ø© ØªØ·Ø§Ø¨Ù‚ Ø§Ù„Ø§Ø³Ù…": 0,
                "Ù†Ø³Ø¨Ø© ØªØ·Ø§Ø¨Ù‚ Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹": 0,
                "ÙƒÙ… ØªÙƒØ±Ø§Ø±Ù‡ ÙÙŠ Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª": 0,
                "Ø§Ù„Ø§Ø³Ù… ÙÙŠ Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª": "â€”",
                "Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨ ÙÙŠ Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª": "â€”",
                "Ø§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹ ÙÙŠ Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª": "â€”"
            }

            if grp_df is not None and not grp_df.empty:
                db_names = grp_df["__norm_name"].tolist()

                # 1) Ø­Ø³Ø§Ø¨ Ù†Ø³Ø¨ Ø§Ù„ØªØ·Ø§Ø¨Ù‚ Ù„Ù„Ø§Ø³Ù…
                scores = process.cdist([norm_name], db_names, scorer=fuzz.token_sort_ratio)[0]

                # 2) ØªØµÙÙŠØ© >= 90%
                HIGH_MATCH_THRESHOLD = 90
                high_match_indices = [j for j, score in enumerate(scores) if score >= HIGH_MATCH_THRESHOLD]

                # 3) Ø§Ù„Ù…Ø±Ø´Ø­ÙˆÙ†
                potential_matches = grp_df.iloc[high_match_indices].copy()
                potential_matches["_name_score"] = scores[high_match_indices]

                match_count = len(potential_matches)
                match_row["ÙƒÙ… ØªÙƒØ±Ø§Ø±Ù‡ ÙÙŠ Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª"] = match_count

                best_match = None
                if match_count == 1:
                    best_match = potential_matches.iloc[0]
                elif match_count > 1:
                    db_centers = potential_matches["__norm_center"].tolist()
                    center_scores = process.cdist([norm_center], db_centers, scorer=fuzz.ratio)[0]
                    potential_matches["_center_score"] = center_scores
                    best_match_idx = potential_matches["_center_score"].argmax()
                    best_match = potential_matches.iloc[best_match_idx]

                # 4) ØªØ¹Ø¨Ø¦Ø© Ø§Ù„Ù†ØªÙŠØ¬Ø©
                if best_match is not None:
                    name_score = best_match["_name_score"]
                    center_score = fuzz.ratio(norm_center, best_match["__norm_center"])
                    match_row.update({
                        "Ù†Ø³Ø¨Ø© ØªØ·Ø§Ø¨Ù‚ Ø§Ù„Ø§Ø³Ù…": round(name_score, 2),
                        "Ù†Ø³Ø¨Ø© ØªØ·Ø§Ø¨Ù‚ Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹": round(center_score, 2),
                        "Ø§Ù„Ø§Ø³Ù… ÙÙŠ Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª": best_match["Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø«Ù„Ø§Ø«ÙŠ"],
                        "Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨ ÙÙŠ Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª": best_match["Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨"],
                        "Ø§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹ ÙÙŠ Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª": best_match["Ø§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹"]
                    })

            results.append(match_row)

            # âœ… ØªØ­Ø¯ÙŠØ« Ø§Ù„ÙˆØ§Ø¬Ù‡Ø©
            if (i + 1) % 50 == 0 or i + 1 == total:
                progress.progress((i + 1) / total)
                log_box.text(f"ğŸ”¹ {i + 1}/{total}: {orig_name[:25]} ...")
                elapsed = time.time() - start
                status.text(f"â±ï¸ ØªÙ…Øª Ù…Ø¹Ø§Ù„Ø¬Ø© {i + 1}/{total} | Ø§Ù„ÙˆÙ‚Øª Ø§Ù„Ù…Ù†Ù‚Ø¶ÙŠ: {elapsed:.1f} Ø«Ø§Ù†ÙŠØ©")

            # ğŸ’¾ Ø­ÙØ¸ Ù…Ø¤Ù‚Øª ÙƒÙ„ 100 Ø³Ø¬Ù„
            if (i + 1) % 100 == 0 or i + 1 == total:
                temp_df = pd.DataFrame(results)
                temp_df.to_excel(temp_path, index=False, sheet_name="Ù†ØªØ§Ø¦Ø¬ Ù…Ø¤Ù‚ØªØ©")

        # ---- Ø§Ù„Ø­ÙØ¸ Ø§Ù„Ù†Ù‡Ø§Ø¦ÙŠ ----
        final_df = pd.DataFrame(results)
        out_file = "Ù†ØªØ§Ø¦Ø¬_Ø§Ù„ØªØ·Ø§Ø¨Ù‚_Ø§Ù„Ù†Ù‡Ø§Ø¦ÙŠØ©.xlsx"
        final_df.to_excel(out_file, index=False)

        st.success(f"âœ… ØªÙ… Ø§ÙƒØªÙ…Ø§Ù„ Ø§Ù„Ø¨Ø­Ø« ÙÙŠ {time.time() - start:.1f} Ø«Ø§Ù†ÙŠØ©")

        # ---- Ø£Ø²Ø±Ø§Ø± Ø§Ù„ØªØ­Ù…ÙŠÙ„ ----
        with open(out_file, "rb") as f1:
            st.download_button("â¬‡ï¸ ØªØ­Ù…ÙŠÙ„ Ø§Ù„Ù†ØªØ§Ø¦Ø¬ Ø§Ù„Ù†Ù‡Ø§Ø¦ÙŠØ©", f1, file_name="Ù†ØªØ§Ø¦Ø¬_Ø§Ù„ØªØ·Ø§Ø¨Ù‚_Ø§Ù„Ù†Ù‡Ø§Ø¦ÙŠØ©.xlsx")

        with open(temp_path, "rb") as f2:
            st.download_button("â¬‡ï¸ ØªØ­Ù…ÙŠÙ„ Ø§Ù„Ù†Ø³Ø®Ø© Ø§Ù„Ø§Ø­ØªÙŠØ§Ø·ÙŠØ© Ø§Ù„Ù…Ø¤Ù‚ØªØ©", f2, file_name="Ù†ØªØ§Ø¦Ø¬_Ù…Ø¤Ù‚ØªØ©.xlsx")

# ----------------------------------------------------------------------------- #
# 5) ğŸ“¦ Ø¹Ø¯Ù‘ Ø§Ù„Ø¨Ø·Ø§Ù‚Ø§Øª (Ø£Ø±Ù‚Ø§Ù… 8 Ø®Ø§Ù†Ø§Øª) + Ø¨Ø­Ø« ÙÙŠ Ø§Ù„Ù‚Ø§Ø¹Ø¯Ø© + Ù‚Ø§Ø¦Ù…Ø© Ø§Ù„Ø£Ø±Ù‚Ø§Ù… ØºÙŠØ± Ø§Ù„Ù…ÙˆØ¬ÙˆØ¯Ø©
# ----------------------------------------------------------------------------- #
with tab_count:
    st.subheader("ğŸ“¦ Ø¹Ø¯Ù‘ Ø§Ù„Ø¨Ø·Ø§Ù‚Ø§Øª (Ø£Ø±Ù‚Ø§Ù… 8 Ø®Ø§Ù†Ø§Øª) â€” Ø¨Ø­Ø« ÙÙŠ Ø§Ù„Ù‚Ø§Ø¹Ø¯Ø© + Ø§Ù„Ø£Ø±Ù‚Ø§Ù… ØºÙŠØ± Ø§Ù„Ù…ÙˆØ¬ÙˆØ¯Ø©")

    imgs_count = st.file_uploader(
        "ğŸ“¤ Ø§Ø±ÙØ¹ ØµÙˆØ± Ø§Ù„ØµÙØ­Ø§Øª (Ù‚Ø¯ ØªØ­ØªÙˆÙŠ Ø£ÙƒØ«Ø± Ù…Ù† Ø¨Ø·Ø§Ù‚Ø©)",
        type=["jpg","jpeg","png"],
        accept_multiple_files=True,
        key="ocr_count"
    )

    if imgs_count and st.button("ğŸš€ Ø¹Ø¯Ù‘ Ø§Ù„Ø¨Ø·Ø§Ù‚Ø§Øª ÙˆØ§Ù„Ø¨Ø­Ø«"):
        client = setup_google_vision()
        if client is None:
            st.error("âŒ Ø®Ø·Ø£ ÙÙŠ Ø¥Ø¹Ø¯Ø§Ø¯ Google Vision.")
        else:
            all_numbers, number_to_files, details = [], {}, []

            for img in imgs_count:
                try:
                    content = img.read()
                    image = vision.Image(content=content)
                    response = client.text_detection(image=image)
                    texts = response.text_annotations
                    full_text = texts[0].description if texts else ""

                    # Ø§Ø³ØªØ®Ø±Ø§Ø¬ Ø£Ø±Ù‚Ø§Ù… Ù…ÙƒÙˆÙ†Ø© Ù…Ù† 8 Ø®Ø§Ù†Ø§Øª ÙÙ‚Ø·
                    found_numbers = re.findall(r"\b\d{8}\b", full_text)
                    for n in found_numbers:
                        all_numbers.append(n)
                        number_to_files.setdefault(n, set()).add(img.name)

                    details.append({
                        "Ø§Ø³Ù… Ø§Ù„Ù…Ù„Ù": img.name,
                        "Ø¹Ø¯Ø¯ Ø§Ù„Ø¨Ø·Ø§Ù‚Ø§Øª (Ø£Ø±Ù‚Ø§Ù… 8 Ø®Ø§Ù†Ø§Øª)": len(found_numbers),
                        "Ø§Ù„Ø£Ø±Ù‚Ø§Ù… Ø§Ù„Ù…ÙƒØªØ´ÙØ© (Ø£Ø±Ù‚Ø§Ù… 8 Ø®Ø§Ù†Ø§Øª ÙÙ‚Ø·)": ", ".join(found_numbers) if found_numbers else "Ù„Ø§ ÙŠÙˆØ¬Ø¯"
                    })

                except Exception as e:
                    st.warning(f"âš ï¸ Ø®Ø·Ø£ Ø£Ø«Ù†Ø§Ø¡ Ù…Ø¹Ø§Ù„Ø¬Ø© ØµÙˆØ±Ø© {img.name}: {e}")

            total_cards = len(all_numbers)
            unique_numbers = sorted(list(set(all_numbers)))

            st.success("âœ… ØªÙ… Ø§Ù„Ø§Ø³ØªØ®Ø±Ø§Ø¬ Ø§Ù„Ø£ÙˆÙ„ÙŠ Ù„Ù„Ø£Ø±Ù‚Ø§Ù…")

            # ----------------- Ø¨Ø­Ø« ÙÙŠ Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª Ø¹Ù† Ø§Ù„Ø£Ø±Ù‚Ø§Ù… Ø§Ù„Ù…ÙˆØ¬ÙˆØ¯Ø© -----------------
            found_df = pd.DataFrame()
            missing_list = []
            if unique_numbers:
                try:
                    conn = get_conn()
                    placeholders = ",".join(["%s"] * len(unique_numbers))
                    query = f"""
                        SELECT "Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨","Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø«Ù„Ø§Ø«ÙŠ","Ø§Ù„Ø¬Ù†Ø³","Ù‡Ø§ØªÙ","Ø±Ù‚Ù… Ø§Ù„Ø¹Ø§Ø¦Ù„Ø©",
                               "Ø§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹","Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹",
                               "Ø§Ù„Ù…Ø¯ÙŠÙ†Ø©","Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„ØªØ³Ø¬ÙŠÙ„","Ø§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„ØªØ³Ø¬ÙŠÙ„","ØªØ§Ø±ÙŠØ® Ø§Ù„Ù…ÙŠÙ„Ø§Ø¯"
                        FROM "Bagdad" WHERE "Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨" IN ({placeholders})
                    """
                    found_df = pd.read_sql_query(query, conn, params=unique_numbers)
                    conn.close()

                    if not found_df.empty:
                        found_df = found_df.rename(columns={
                            "Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨": "Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨",
                            "Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø«Ù„Ø§Ø«ÙŠ": "Ø§Ù„Ø§Ø³Ù…",
                            "Ø§Ù„Ø¬Ù†Ø³": "Ø§Ù„Ø¬Ù†Ø³",
                            "Ù‡Ø§ØªÙ": "Ø±Ù‚Ù… Ø§Ù„Ù‡Ø§ØªÙ",
                            "Ø±Ù‚Ù… Ø§Ù„Ø¹Ø§Ø¦Ù„Ø©": "Ø±Ù‚Ù… Ø§Ù„Ø¹Ø§Ø¦Ù„Ø©",
                            "Ø§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹": "Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹",
                            "Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹": "Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹",
                            "Ø§Ù„Ù…Ø¯ÙŠÙ†Ø©": "Ø§Ù„Ù…Ø¯ÙŠÙ†Ø©",
                            "Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„ØªØ³Ø¬ÙŠÙ„": "Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„ØªØ³Ø¬ÙŠÙ„",
                            "Ø§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„ØªØ³Ø¬ÙŠÙ„": "Ø§Ø³Ù… Ù…Ø±ÙƒØ² Ø§Ù„ØªØ³Ø¬ÙŠÙ„",
                            "ØªØ§Ø±ÙŠØ® Ø§Ù„Ù…ÙŠÙ„Ø§Ø¯": "ØªØ§Ø±ÙŠØ® Ø§Ù„Ù…ÙŠÙ„Ø§Ø¯"
                        })
                        found_df["Ø§Ù„Ø¬Ù†Ø³"] = found_df["Ø§Ù„Ø¬Ù†Ø³"].apply(map_gender)

                        # ğŸ§© Ø¥Ø¶Ø§ÙØ© Ù†ÙØ³ Ø§Ù„Ø£Ø¹Ù…Ø¯Ø© Ù…Ø«Ù„ ØªØ¨ÙˆÙŠØ¨ ğŸ“‚ Ø±ÙØ¹ Ù…Ù„Ù Excel
                        found_df["Ø±Ù‚Ù… Ø§Ù„Ù…Ù†Ø¯ÙˆØ¨ Ø§Ù„Ø±Ø¦ÙŠØ³ÙŠ"] = ""
                        found_df["Ø§Ù„Ø­Ø§Ù„Ø©"] = 0
                        found_df["Ù…Ù„Ø§Ø­Ø¸Ø©"] = ""
                        found_df["Ø±Ù‚Ù… Ø§Ù„Ù…Ø­Ø·Ø©"] = 1

                        # ØªØ±ØªÙŠØ¨ Ø§Ù„Ø£Ø¹Ù…Ø¯Ø©
                        found_df = found_df[[
                            "Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨","Ø§Ù„Ø§Ø³Ù…","Ø§Ù„Ø¬Ù†Ø³","Ø±Ù‚Ù… Ø§Ù„Ù‡Ø§ØªÙ",
                            "Ø±Ù‚Ù… Ø§Ù„Ø¹Ø§Ø¦Ù„Ø©","Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹","Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹","Ø±Ù‚Ù… Ø§Ù„Ù…Ø­Ø·Ø©",
                            "Ø±Ù‚Ù… Ø§Ù„Ù…Ù†Ø¯ÙˆØ¨ Ø§Ù„Ø±Ø¦ÙŠØ³ÙŠ","Ø§Ù„Ø­Ø§Ù„Ø©","Ù…Ù„Ø§Ø­Ø¸Ø©"
                        ]]

                    found_numbers_in_db = set(found_df["Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨"].astype(str).tolist()) if not found_df.empty else set()
                    for n in unique_numbers:
                        if n not in found_numbers_in_db:
                            files = sorted(list(number_to_files.get(n, [])))
                            missing_list.append({"Ø±Ù‚Ù…_Ø§Ù„Ù†Ø§Ø®Ø¨": n, "Ø§Ù„Ù…ØµØ¯Ø±(Ø§Ù„ØµÙˆØ±)": ", ".join(files)})

                except Exception as e:
                    st.error(f"âŒ Ø®Ø·Ø£ Ø£Ø«Ù†Ø§Ø¡ Ø§Ù„Ø¨Ø­Ø« ÙÙŠ Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª: {e}")
            else:
                st.info("â„¹ï¸ Ù„Ù… ÙŠØªÙ… Ø§Ù„Ø¹Ø«ÙˆØ± Ø¹Ù„Ù‰ Ø£ÙŠ Ø£Ø±Ù‚Ø§Ù… Ù…ÙƒÙˆÙ‘Ù†Ø© Ù…Ù† 8 Ø®Ø§Ù†Ø§Øª ÙÙŠ Ø§Ù„ØµÙˆØ± Ø§Ù„Ù…Ø±ÙÙˆØ¹Ø©.")

            # ----------------- Ø¹Ø±Ø¶ Ø§Ù„Ù†ØªØ§Ø¦Ø¬ Ù„Ù„Ù…Ø³ØªØ®Ø¯Ù… -----------------
            st.markdown("### ğŸ“Š Ù…Ù„Ø®Øµ Ø§Ù„Ø§Ø³ØªØ®Ø±Ø§Ø¬")
            c1, c2, c3 = st.columns(3)
            c1.metric("Ø¥Ø¬Ù…Ø§Ù„ÙŠ Ø§Ù„Ø£Ø±Ù‚Ø§Ù… (Ù…Ø¹ Ø§Ù„ØªÙƒØ±Ø§Ø±)", total_cards)
            c2.metric("Ø¥Ø¬Ù…Ø§Ù„ÙŠ Ø§Ù„Ø£Ø±Ù‚Ø§Ù… Ø§Ù„ÙØ±ÙŠØ¯Ø© (8 Ø®Ø§Ù†Ø§Øª)", len(unique_numbers))
            c3.metric("Ø¹Ø¯Ø¯ Ø§Ù„ØµÙˆØ± Ø§Ù„Ù…Ø±ÙÙˆØ¹Ø©", len(imgs_count))

            st.markdown("### ğŸ” Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„Ù†Ø§Ø®Ø¨ÙŠÙ† (Ø§Ù„Ù…ÙˆØ¬ÙˆØ¯Ø© ÙÙŠ Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª)")
            if not found_df.empty:
                st.dataframe(found_df, use_container_width=True, height=400)
                out_found = "Ø¨ÙŠØ§Ù†Ø§Øª_Ø§Ù„Ù†Ø§Ø®Ø¨ÙŠÙ†_Ø§Ù„Ù…ÙˆØ¬ÙˆØ¯ÙŠÙ†.xlsx"
                found_df.to_excel(out_found, index=False, engine="openpyxl")
                wb = load_workbook(out_found)
                wb.active.sheet_view.rightToLeft = True
                wb.save(out_found)
                with open(out_found, "rb") as f:
                    st.download_button("â¬‡ï¸ ØªØ­Ù…ÙŠÙ„ Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„Ù†Ø§Ø®Ø¨ÙŠÙ† Ø§Ù„Ù…ÙˆØ¬ÙˆØ¯Ø©", f,
                        file_name="Ø¨ÙŠØ§Ù†Ø§Øª_Ø§Ù„Ù†Ø§Ø®Ø¨ÙŠÙ†_Ø§Ù„Ù…ÙˆØ¬ÙˆØ¯ÙŠÙ†.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                st.warning("âš ï¸ Ù„Ù… ÙŠØªÙ… Ø§Ù„Ø¹Ø«ÙˆØ± Ø¹Ù„Ù‰ Ø£ÙŠ Ù…Ø·Ø§Ø¨Ù‚Ø§Øª ÙÙŠ Ø§Ù„Ù‚Ø§Ø¹Ø¯Ø©.")

            st.markdown("### âŒ Ø§Ù„Ø£Ø±Ù‚Ø§Ù… ØºÙŠØ± Ø§Ù„Ù…ÙˆØ¬ÙˆØ¯Ø© ÙÙŠ Ø§Ù„Ù‚Ø§Ø¹Ø¯Ø© (Ù…Ø¹ Ø§Ø³Ù… Ø§Ù„ØµÙˆØ±Ø©)")
            if missing_list:
                missing_df = pd.DataFrame(missing_list)
                st.dataframe(missing_df, use_container_width=True)
                miss_file = "missing_numbers_with_files.xlsx"
                missing_df.to_excel(miss_file, index=False, engine="openpyxl")
                with open(miss_file, "rb") as f:
                    st.download_button("â¬‡ï¸ ØªØ­Ù…ÙŠÙ„ Ø§Ù„Ø£Ø±Ù‚Ø§Ù… ØºÙŠØ± Ø§Ù„Ù…ÙˆØ¬ÙˆØ¯Ø© Ù…Ø¹ Ø§Ù„Ù…ØµØ¯Ø±", f,
                        file_name="Ø§Ù„Ø£Ø±Ù‚Ø§Ù…_ØºÙŠØ±_Ø§Ù„Ù…ÙˆØ¬ÙˆØ¯Ø©_Ù…Ø¹_Ø§Ù„Ù…ØµØ¯Ø±.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                st.success("âœ… Ù„Ø§ ØªÙˆØ¬Ø¯ Ø£Ø±Ù‚Ø§Ù… Ù…ÙÙ‚ÙˆØ¯Ø© (ÙƒÙ„ Ø§Ù„Ø£Ø±Ù‚Ø§Ù… Ø§Ù„Ù…ÙˆØ¬ÙˆØ¯Ø© ØªÙ… Ø¥ÙŠØ¬Ø§Ø¯Ù‡Ø§ ÙÙŠ Ù‚Ø§Ø¹Ø¯Ø© Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª).")

# ----------------------------------------------------------------------------- #
# 6) ğŸ§¾ Ø§Ù„ØªØ­Ù‚Ù‚ Ù…Ù† ØµØ­Ø© Ø§Ù„Ù…Ø¹Ù„ÙˆÙ…Ø§Øª (Ø¨ÙˆØ§Ø³Ø·Ø© Ø¨Ø§Ø³Ù…)
# ----------------------------------------------------------------------------- #
with tab_check:
    st.subheader("ğŸ§¾ Ø§Ù„ØªØ­Ù‚Ù‚ Ù…Ù† ØµØ­Ø© Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„Ù†Ø§Ø®Ø¨ÙŠÙ† (Ø¨ÙˆØ§Ø³Ø·Ø© Ø¨Ø§Ø³Ù… âš¡ Ø³Ø±ÙŠØ¹)")

    st.markdown("""
    **ğŸ“‹ Ø§Ù„ØªØ¹Ù„ÙŠÙ…Ø§Øª:**
    - Ø§Ù„Ù…Ù„Ù ÙŠØ¬Ø¨ Ø£Ù† ÙŠØ­ØªÙˆÙŠ Ø§Ù„Ø£Ø¹Ù…Ø¯Ø© Ø§Ù„ØªØ§Ù„ÙŠØ©:
      1. Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨  
      2. Ø§Ù„Ø§Ø³Ù…  
      3. Ø±Ù‚Ù… Ø§Ù„Ø¹Ø§Ø¦Ù„Ø©  
      4. Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹
    """)

    uploaded_check = st.file_uploader("ğŸ“¤ Ø§Ø±ÙØ¹ Ù…Ù„Ù Excel Ù„Ù„ØªØ­Ù‚Ù‚", type=["xlsx"], key="check_file")

    if uploaded_check and st.button("ğŸš€ Ø¨Ø¯Ø¡ Ø§Ù„ØªØ­Ù‚Ù‚ Ø§Ù„Ø³Ø±ÙŠØ¹ Ø¨ÙˆØ§Ø³Ø·Ø© Ø¨Ø§Ø³Ù…"):
        try:
            df_check = pd.read_excel(uploaded_check, engine="openpyxl")

            required_cols = ["Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨", "Ø§Ù„Ø§Ø³Ù…", "Ø±Ù‚Ù… Ø§Ù„Ø¹Ø§Ø¦Ù„Ø©", "Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹"]
            missing_cols = [c for c in required_cols if c not in df_check.columns]

            if missing_cols:
                st.error(f"âŒ Ø§Ù„Ù…Ù„Ù Ù†Ø§Ù‚Øµ Ø§Ù„Ø£Ø¹Ù…Ø¯Ø© Ø§Ù„ØªØ§Ù„ÙŠØ©: {', '.join(missing_cols)}")
            else:
                # Ø¥Ø¸Ù‡Ø§Ø± Ø´Ø±ÙŠØ· Ø§Ù„ØªÙ‚Ø¯Ù… Ø¨Ø§Ø³Ù…
                progress_bar = st.progress(0, text="ğŸ¤– Ø¨Ø§Ø³Ù… ÙŠØ­Ø¶Ù‘Ø± Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª...")
                total_steps = 4

                # Ø§Ù„Ø®Ø·ÙˆØ© 1ï¸âƒ£ - ØªØ­Ù…ÙŠÙ„ Ø£Ø±Ù‚Ø§Ù… Ø§Ù„Ù†Ø§Ø®Ø¨ÙŠÙ† Ù…Ù† Ø§Ù„Ù…Ù„Ù
                voter_list = df_check["Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨"].astype(str).tolist()
                progress_bar.progress(1/total_steps, text="ğŸ“¥ ØªØ­Ù…ÙŠÙ„ Ø£Ø±Ù‚Ø§Ù… Ø§Ù„Ù†Ø§Ø®Ø¨ÙŠÙ†...")

                # Ø§Ù„Ø®Ø·ÙˆØ© 2ï¸âƒ£ - Ø¬Ù„Ø¨ ÙƒÙ„ Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª Ù…Ù† Ø§Ù„Ù‚Ø§Ø¹Ø¯Ø© Ø¯ÙØ¹Ø© ÙˆØ§Ø­Ø¯Ø©
                conn = get_conn()
                placeholders = ",".join(["%s"] * len(voter_list))
                query = f"""
                    SELECT "Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨","Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø«Ù„Ø§Ø«ÙŠ","Ø±Ù‚Ù… Ø§Ù„Ø¹Ø§Ø¦Ù„Ø©","Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹"
                    FROM "Bagdad"
                    WHERE "Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨" IN ({placeholders})
                """
                df_db = pd.read_sql_query(query, conn, params=voter_list)
                conn.close()
                progress_bar.progress(2/total_steps, text="ğŸ’¾ ØªÙ… Ø¬Ù„Ø¨ Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª Ù…Ù† Ø§Ù„Ù‚Ø§Ø¹Ø¯Ø©...")

                # Ø§Ù„Ø®Ø·ÙˆØ© 3ï¸âƒ£ - Ø¯Ù…Ø¬ Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª Ø­Ø³Ø¨ Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨
                merged = pd.merge(
                    df_check.astype(str),
                    df_db.astype(str),
                    on="Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨",
                    how="left",
                    suffixes=("_Ø§Ù„Ù…Ø¯Ø®Ù„", "_Ø§Ù„Ù‚Ø§Ø¹Ø¯Ø©")
                )

                progress_bar.progress(3/total_steps, text="ğŸ§  Ø¨Ø§Ø³Ù… ÙŠÙ‚Ø§Ø±Ù† Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª...")

                # Ø§Ù„Ø®Ø·ÙˆØ© 4ï¸âƒ£ - Ø§Ù„ØªØ­Ù‚Ù‚ Ù…Ù† Ø§Ù„ØªØ·Ø§Ø¨Ù‚Ø§Øª
                def match(a, b): return "âœ…" if a == b else "âŒ"

                merged["ØªØ·Ø§Ø¨Ù‚ Ø§Ù„Ø§Ø³Ù…"] = merged.apply(lambda r: match(r["Ø§Ù„Ø§Ø³Ù…"], r["Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø«Ù„Ø§Ø«ÙŠ"]), axis=1)
                merged["ØªØ·Ø§Ø¨Ù‚ Ø±Ù‚Ù… Ø§Ù„Ø¹Ø§Ø¦Ù„Ø©"] = merged.apply(lambda r: match(r["Ø±Ù‚Ù… Ø§Ù„Ø¹Ø§Ø¦Ù„Ø©_Ø§Ù„Ù…Ø¯Ø®Ù„"], r["Ø±Ù‚Ù… Ø§Ù„Ø¹Ø§Ø¦Ù„Ø©_Ø§Ù„Ù‚Ø§Ø¹Ø¯Ø©"]), axis=1)
                merged["ØªØ·Ø§Ø¨Ù‚ Ø§Ù„Ù…Ø±ÙƒØ²"] = merged.apply(lambda r: match(r["Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹_Ø§Ù„Ù…Ø¯Ø®Ù„"], r["Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹_Ø§Ù„Ù‚Ø§Ø¹Ø¯Ø©"]), axis=1)

                def overall(row):
                    if pd.isna(row["Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø«Ù„Ø§Ø«ÙŠ"]):
                        return "âŒ ØºÙŠØ± Ù…ÙˆØ¬ÙˆØ¯ ÙÙŠ Ø§Ù„Ù‚Ø§Ø¹Ø¯Ø©"
                    elif all(row[x] == "âœ…" for x in ["ØªØ·Ø§Ø¨Ù‚ Ø§Ù„Ø§Ø³Ù…", "ØªØ·Ø§Ø¨Ù‚ Ø±Ù‚Ù… Ø§Ù„Ø¹Ø§Ø¦Ù„Ø©", "ØªØ·Ø§Ø¨Ù‚ Ø§Ù„Ù…Ø±ÙƒØ²"]):
                        return "âœ… Ù…Ø·Ø§Ø¨Ù‚"
                    else:
                        return "âš ï¸ Ø§Ø®ØªÙ„Ø§Ù"

                merged["Ø§Ù„Ù†ØªÙŠØ¬Ø© Ø§Ù„Ù†Ù‡Ø§Ø¦ÙŠØ©"] = merged.apply(overall, axis=1)

                progress_bar.progress(1.0, text="âœ… ØªÙ… Ø§Ù„ØªØ­Ù‚Ù‚ Ø¨ÙˆØ§Ø³Ø·Ø© Ø¨Ø§Ø³Ù… Ø¨Ø³Ø±Ø¹Ø© âš¡")

                # Ø¹Ø±Ø¶ Ø§Ù„Ù†ØªØ§Ø¦Ø¬
                st.dataframe(merged[[
                    "Ø±Ù‚Ù… Ø§Ù„Ù†Ø§Ø®Ø¨",
                    "Ø§Ù„Ø§Ø³Ù…",
                    "Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø«Ù„Ø§Ø«ÙŠ",
                    "ØªØ·Ø§Ø¨Ù‚ Ø§Ù„Ø§Ø³Ù…",
                    "Ø±Ù‚Ù… Ø§Ù„Ø¹Ø§Ø¦Ù„Ø©_Ø§Ù„Ù…Ø¯Ø®Ù„",
                    "Ø±Ù‚Ù… Ø§Ù„Ø¹Ø§Ø¦Ù„Ø©_Ø§Ù„Ù‚Ø§Ø¹Ø¯Ø©",
                    "ØªØ·Ø§Ø¨Ù‚ Ø±Ù‚Ù… Ø§Ù„Ø¹Ø§Ø¦Ù„Ø©",
                    "Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹_Ø§Ù„Ù…Ø¯Ø®Ù„",
                    "Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹_Ø§Ù„Ù‚Ø§Ø¹Ø¯Ø©",
                    "ØªØ·Ø§Ø¨Ù‚ Ø§Ù„Ù…Ø±ÙƒØ²",
                    "Ø§Ù„Ù†ØªÙŠØ¬Ø© Ø§Ù„Ù†Ù‡Ø§Ø¦ÙŠØ©"
                ]], use_container_width=True, height=450)

                # ØªØ­Ù…ÙŠÙ„ Ù…Ù„Ù Ø§Ù„Ù†ØªØ§Ø¦Ø¬
                out_file = "Ù†ØªØ§Ø¦Ø¬_Ø§Ù„ØªØ­Ù‚Ù‚_Ø§Ù„Ø³Ø±ÙŠØ¹.xlsx"
                merged.to_excel(out_file, index=False, engine="openpyxl")
                with open(out_file, "rb") as f:
                    st.download_button("â¬‡ï¸ ØªØ­Ù…ÙŠÙ„ Ù†ØªØ§Ø¦Ø¬ Ø§Ù„ØªØ­Ù‚Ù‚ Ø§Ù„Ø³Ø±ÙŠØ¹", f,
                        file_name="Ù†ØªØ§Ø¦Ø¬_Ø§Ù„ØªØ­Ù‚Ù‚_Ø§Ù„Ø³Ø±ÙŠØ¹.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        except Exception as e:
            st.error(f"âŒ Ø®Ø·Ø£ Ø£Ø«Ù†Ø§Ø¡ Ø§Ù„ØªØ­Ù‚Ù‚: {e}")

# ----------------------------------------------------------------------------- #
# 8) ğŸ§© ØªØ­Ù„ÙŠÙ„ Ù…Ø®ØµØµ (Group & Count)
# ----------------------------------------------------------------------------- #
st.markdown("---")
st.subheader("ğŸ§© ØªØ­Ù„ÙŠÙ„ Ù…Ø®ØµØµ (Group & Count)")

st.markdown("""
**ğŸ“‹ Ø§Ù„ØªØ¹Ù„ÙŠÙ…Ø§Øª:**
- Ø§Ø±ÙØ¹ Ù…Ù„Ù Excel ÙŠØ­ØªÙˆÙŠ Ø§Ù„Ø£Ø¹Ù…Ø¯Ø© Ø§Ù„Ù…Ø·Ù„ÙˆØ¨Ø©.  
- Ø§Ø®ØªØ± Ø§Ù„Ø£Ø¹Ù…Ø¯Ø© Ø§Ù„ØªÙŠ ØªØ±ÙŠØ¯ ØªØ¬Ù…ÙŠØ¹ Ø§Ù„Ù†ØªØ§Ø¦Ø¬ Ø¨Ù†Ø§Ø¡Ù‹ Ø¹Ù„ÙŠÙ‡Ø§ (Ù…Ø«Ù„Ø§Ù‹: *Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„Ø§Ù‚ØªØ±Ø§Ø¹ + Ø±Ù‚Ù… Ù…Ø±ÙƒØ² Ø§Ù„ØªØ³Ø¬ÙŠÙ„*).  
- Ø§Ø®ØªØ± Ø§Ù„Ø¹Ù…ÙˆØ¯ Ø§Ù„Ø°ÙŠ ØªØ±ÙŠØ¯ Ø­Ø³Ø§Ø¨ Ø¹Ø¯Ø¯ ØªÙƒØ±Ø§Ø±Ø§ØªÙ‡ (COUNT).  
- Ø¨Ø§Ø³Ù… Ø³ÙŠÙØ¸Ù‡Ø± Ù„Ùƒ Ø¹Ø¯Ø¯ Ø§Ù„ØµÙÙˆÙ (Ø§Ù„Ù†Ø§Ø®Ø¨ÙŠÙ† Ù…Ø«Ù„Ø§Ù‹) Ø¶Ù…Ù† ÙƒÙ„ Ù…Ø¬Ù…ÙˆØ¹Ø© ğŸ‘‡
""")

uploaded_group = st.file_uploader("ğŸ“¤ Ø§Ø±ÙØ¹ Ù…Ù„Ù Excel Ù„Ù„ØªØ­Ù„ÙŠÙ„ Ø§Ù„Ù…Ø®ØµØµ", type=["xlsx"], key="group_file")

if uploaded_group:
    try:
        df = pd.read_excel(uploaded_group, engine="openpyxl")
        st.success(f"âœ… ØªÙ… ØªØ­Ù…ÙŠÙ„ Ø§Ù„Ù…Ù„Ù ({len(df)} ØµÙ)")

        st.markdown("### ğŸ§± Ø§Ù„Ø£Ø¹Ù…Ø¯Ø© Ø§Ù„Ù…ØªÙˆÙØ±Ø©:")
        st.write(list(df.columns))

        group_cols = st.multiselect("ğŸ“Š Ø§Ø®ØªØ± Ø§Ù„Ø£Ø¹Ù…Ø¯Ø© Ù„Ù„ØªØ¬Ù…ÙŠØ¹ (Group By):", options=df.columns)
        count_col = st.selectbox("ğŸ”¢ Ø§Ø®ØªØ± Ø§Ù„Ø¹Ù…ÙˆØ¯ Ø§Ù„Ù…Ø±Ø§Ø¯ Ø¹Ø¯Ù‡ (COUNT):", options=df.columns)

        if group_cols and count_col and st.button("ğŸš€ ØªÙ†ÙÙŠØ° Ø§Ù„ØªØ­Ù„ÙŠÙ„ Ø§Ù„Ù…Ø®ØµØµ"):
            progress = st.progress(0, text="ğŸ¤– Ø¨Ø§Ø³Ù… ÙŠØ­Ù„Ù„ Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª...")
            total_steps = 3

            # Ø§Ù„Ø®Ø·ÙˆØ© 1ï¸âƒ£ - ØªØ¬Ù‡ÙŠØ² Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª
            progress.progress(1/total_steps, text="ğŸ§® ØªØ¬Ù…ÙŠØ¹ Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª...")

            # Ø§Ù„Ø®Ø·ÙˆØ© 2ï¸âƒ£ - Ø­Ø³Ø§Ø¨ Ø¹Ø¯Ø¯ Ø§Ù„ØµÙÙˆÙ Ø­Ø³Ø¨ Ø§Ù„Ø£Ø¹Ù…Ø¯Ø© Ø§Ù„Ù…Ø­Ø¯Ø¯Ø©
            grouped = df.groupby(group_cols)[count_col].count().reset_index()
            grouped = grouped.rename(columns={count_col: "Ø¹Ø¯Ø¯ Ø§Ù„ØµÙÙˆÙ"})

            progress.progress(2/total_steps, text="ğŸ“Š ØªØ¬Ù‡ÙŠØ² Ø§Ù„Ù†ØªØ§Ø¦Ø¬...")

            # Ø§Ù„Ø®Ø·ÙˆØ© 3ï¸âƒ£ - Ø¹Ø±Ø¶ ÙˆØªØ­Ù…ÙŠÙ„ Ø§Ù„Ù†ØªØ§Ø¦Ø¬
            st.dataframe(grouped, use_container_width=True, height=450)

            # Ø²Ø± ØªØ­Ù…ÙŠÙ„ Ø§Ù„Ù†ØªØ§Ø¦Ø¬
            out_file = "Ù†ØªØ§Ø¦Ø¬_ØªØ­Ù„ÙŠÙ„_Ù…Ø®ØµØµ.xlsx"
            grouped.to_excel(out_file, index=False, engine="openpyxl")
            with open(out_file, "rb") as f:
                st.download_button("â¬‡ï¸ ØªØ­Ù…ÙŠÙ„ Ø§Ù„Ù†ØªØ§Ø¦Ø¬ (Excel)", f,
                    file_name="Ù†ØªØ§Ø¦Ø¬_ØªØ­Ù„ÙŠÙ„_Ù…Ø®ØµØµ.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            progress.progress(1.0, text="âœ… ØªÙ… Ø§Ù„ØªØ­Ù„ÙŠÙ„ Ø¨Ù†Ø¬Ø§Ø­ Ø¨ÙˆØ§Ø³Ø·Ø© Ø¨Ø§Ø³Ù…!")
    except Exception as e:
        st.error(f"âŒ Ø­Ø¯Ø« Ø®Ø·Ø£ Ø£Ø«Ù†Ø§Ø¡ Ø§Ù„ØªØ­Ù„ÙŠÙ„: {e}")
